#! /usr/bin/python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('/home/jesse/Dropbox/computers/Automate the Boring Stuff with Python - Practical Programming for Total Beginners - 1st Edition (2015) (Pdf, Epub & Mobi) Gooner/automate_online-materials/produceSales.xlsx')

sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices 
PRICE_UPDATES = {'Garlic': 4.07, 'Celery': 2.19, 'Lemon': 5.27}

# TODO: Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row): 	#skip first row
	produceName = sheet.cell(row=rowNum, column=1).value
	if produceName in PRICE_UPDATES:
		sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
